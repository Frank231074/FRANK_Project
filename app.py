import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# エクセルファイルの読み込み
wb = openpyxl.load_workbook("files/製造マスタサンプル.xlsx",data_only=True)
ws = wb.active

values = list(ws.values)
lastrow = len()
print(file_path)
