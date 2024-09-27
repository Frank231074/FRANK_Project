import openpyxl
from openpyxl.styles import PatternFill

# Excelファイルをロード
wb = openpyxl.load_workbook("files/製造マスタサンプル.xlsx")

# アクティブなシートを選択
ws = wb.active

# B列、E列、J列のすべてのセルの値をリストに保存
b_column_values = []
e_column_values = []
j_column_values = []

# B列の値を取得
# 2行目以降を指定
for cell in ws["B"][2:]:
    if cell.value is not None:  # Noneじゃない場合のみを追加
        b_column_values.append(cell.value)


# E列の値を取得
for cell in ws["E"][2:]:
    if cell.value is not None:
        e_column_values.append(cell.value)

# J列の値を取得
for cell in ws["J"][2:]:
    if cell.value is not None:
        j_column_values.append(cell.value)

#　B列とJ列の値がマッチしている値を取得
matched_list_B = []
for b_column in b_column_values:
    for j_column in j_column_values:
        if b_column == j_column:
            matched_list_B.append(b_column)

#　E列とJ列の値がマッチしている値を取得
matched_list_E = []
for e_column in e_column_values:
    for j_column in j_column_values:
        if e_column == j_column:
            matched_list_E.append(e_column)

# セルの色を設定する
fill_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# B列の2行目から順にセルを一つずつ取り出す
# 取り出したセルの値が、matched_listに含まれているか調べる
# 含まれている場合、そのセルの塗りつぶしの色をfill_colorに変更する
# 全てのセルに対して上記処理を繰り返す
for row, cell in enumerate(ws["B"][1:], start=2):
    if cell.value in matched_list_B:
        ws[f"B{row}"].fill = fill_color

# E列も同様の処理を行う
for row, cell in enumerate(ws["E"][1:], start=2):
    if cell.value in matched_list_E:
        ws[f"E{row}"].fill = fill_color


# 名前を付けて保存
wb.save("result001.xlsx")

# print(b_column_values)
# print(e_column_values)
# print(j_column_values)
# print(matched_list)
# print(e_column_values)
# print(j_column_values)
