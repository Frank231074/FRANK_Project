import openpyxl
from openpyxl.styles import PatternFill

# Excelファイルをロード
wb = openpyxl.load_workbook("files/製造マスタサンプル２.xlsx")

# アクティブなシートを選択
ws_master = wb["製造マスタ"]
ws_search = wb["検索シート"]
search_code = ws_search["A1"].value

# ws_master_B列、E列、ws_seach_B列のすべてのセルの値をリストに保存
ws_master_b_column_values = []
ws_master_e_column_values = []
ws_search_b_column_values = []

# ws_master_B列の値を取得
# 2行目以降を指定
for cell in ws_master["B"][2:]:
    if cell.value is not None:  # Noneじゃない場合のみを追加
        ws_master_b_column_values.append(cell.value)


# ws_master_E列の値を取得
for cell in ws_master["E"][2:]:
    if cell.value is not None:
        ws_master_e_column_values.append(cell.value)

# ws_search_b列の値を取得
for cell in ws_search["B"][2:]:
    if cell.value is not None:
        ws_search_b_column_values.append(cell.value)

# 　master_B列とsearch_B列の値がマッチしている値を取得
matched_list_B = []
for master_b_column in ws_master_b_column_values:
    for search_b_column in ws_search_b_column_values:
        if master_b_column == search_b_column:
            matched_list_B.append(master_b_column)

# 　master_E列とsearch_B列がマッチしている値を取得
matched_list_E = []
for master_e_column in ws_master_e_column_values:
    for search_b_column in ws_search_b_column_values:
        if master_e_column == search_b_column:
            matched_list_E.append(master_e_column)

# セルの色を設定する
fill_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# B列の2行目から順にセルを一つずつ取り出す
# 取り出したセルの値が、matched_listに含まれているか調べる
# 含まれている場合、そのセルの塗りつぶしの色をfill_colorに変更する
# 全てのセルに対して上記処理を繰り返す
for row, cell in enumerate(ws_master["B"][1:], start=2):
    if cell.value in matched_list_B:
        ws_master[f"B{row}"].fill = fill_color

# E列も同様の処理を行う
for row, cell in enumerate(ws_master["E"][1:], start=2):
    if cell.value in matched_list_E:
        ws_master[f"E{row}"].fill = fill_color


# 名前を付けて保存
output_file = f"result_{search_code}.xlsx"

wb.save(output_file)

# print(b_column_values)
# print(e_column_values)
# print(j_column_values)
# print(matched_list)
# print(e_column_values)
# print(j_column_values)
